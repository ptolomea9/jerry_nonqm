import csv
import os
import uuid
import threading
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, current_app
import config
from models import get_db, query_db

bp = Blueprint("lists", __name__)


def _parse_upload(filepath):
    """Parse a CSV or XLSX file, returning (header, rows) as lists of strings."""
    ext = os.path.splitext(filepath)[1].lower()

    if ext == ".xlsx":
        from openpyxl import load_workbook

        def _cell_to_str(val):
            if val is None:
                return ""
            if isinstance(val, float) and val == int(val):
                return str(int(val))
            return str(val)

        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows()
        header = [_cell_to_str(cell.value) for cell in next(rows_iter)]
        rows = []
        for row in rows_iter:
            rows.append([_cell_to_str(cell.value) for cell in row])
        wb.close()
        return header, rows
    else:
        with open(filepath, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            header = next(reader)
            rows = list(reader)
        return header, rows


def _pick_column_map(header):
    """Auto-detect which column map has more header matches."""
    from import_csv import COLUMN_MAP, XLSX_COLUMN_MAP

    csv_hits = sum(1 for col in COLUMN_MAP if col in header)
    xlsx_hits = sum(1 for col in XLSX_COLUMN_MAP if col in header)

    return XLSX_COLUMN_MAP if xlsx_hits > csv_hits else COLUMN_MAP


@bp.route("/lists")
def index():
    lists = query_db("SELECT * FROM lists ORDER BY created_at DESC")
    return render_template("lists/index.html", lists=lists)


@bp.route("/lists/upload", methods=["POST"])
def upload():
    if "csv_file" not in request.files:
        flash("No file selected", "error")
        return redirect(url_for("lists.index"))

    file = request.files["csv_file"]
    if file.filename == "":
        flash("No file selected", "error")
        return redirect(url_for("lists.index"))

    ext = os.path.splitext(file.filename)[1].lower().lstrip(".")
    if ext not in config.ALLOWED_CSV_EXTENSIONS:
        flash("Only CSV and XLSX files are allowed", "error")
        return redirect(url_for("lists.index"))

    list_name = request.form.get("list_name", "").strip() or file.filename

    # Save file
    os.makedirs(config.UPLOAD_FOLDER_CSV, exist_ok=True)
    stored_name = f"{uuid.uuid4().hex}_{file.filename}"
    filepath = os.path.join(config.UPLOAD_FOLDER_CSV, stored_name)
    file.save(filepath)

    # Parse file (CSV or XLSX)
    try:
        header, rows = _parse_upload(filepath)
    except Exception as e:
        flash(f"Error reading file: {e}", "error")
        return redirect(url_for("lists.index"))

    db = get_db()

    # Auto-detect column map
    active_map = _pick_column_map(header)

    # Determine enrichment status from headers
    has_website = "Company Website" in header
    has_socials = "Facebook" in header
    enrichment_status = "complete" if (has_website and has_socials) else "none"

    # Create list record
    cur = db.execute(
        "INSERT INTO lists (name, filename, row_count, enrichment_status) VALUES (?, ?, ?, ?)",
        (list_name, stored_name, len(rows), enrichment_status),
    )
    list_id = cur.lastrowid

    # Build column index mapping
    col_indices = {}
    for csv_col, db_col in active_map.items():
        if csv_col in header:
            col_indices[header.index(csv_col)] = db_col

    if not col_indices:
        db.commit()
        flash(f"Warning: No matching columns found. Created list with 0 leads.", "error")
        return redirect(url_for("lists.detail", list_id=list_id))

    # Find which column maps to nmlsid
    nmlsid_idx = None
    for idx, db_col in col_indices.items():
        if db_col == "nmlsid":
            nmlsid_idx = idx
            break

    if nmlsid_idx is None:
        db.commit()
        flash("Error: No NMLSID column found in file", "error")
        return redirect(url_for("lists.detail", list_id=list_id))

    sorted_indices = sorted(col_indices.keys())
    db_columns = [col_indices[idx] for idx in sorted_indices]
    placeholders = ", ".join(["?"] * len(db_columns))
    col_names = ", ".join(db_columns)
    update_cols = [c for c in db_columns if c != "nmlsid"]
    update_set = ", ".join(f"{c} = excluded.{c}" for c in update_cols)

    insert_sql = f"""
        INSERT INTO leads ({col_names})
        VALUES ({placeholders})
        ON CONFLICT(nmlsid) DO UPDATE SET {update_set}
    """

    lead_ids = []

    for row in rows:
        # Skip rows with blank/missing nmlsid
        nmlsid_val = row[nmlsid_idx] if nmlsid_idx < len(row) else ""
        if not nmlsid_val or nmlsid_val.strip() == "" or nmlsid_val == "None":
            continue

        values = []
        for idx in sorted_indices:
            val = row[idx] if idx < len(row) else ""
            # Clean "None" string values from XLSX
            if val == "None":
                val = ""
            values.append(val)

        try:
            db.execute(insert_sql, values)
            lead_row = db.execute("SELECT id FROM leads WHERE nmlsid = ?", (nmlsid_val.strip(),)).fetchone()
            if lead_row:
                lead_ids.append(lead_row["id"])
        except Exception as e:
            print(f"  Error importing row: {e}")

    # Update row_count with actual leads linked
    db.execute("UPDATE lists SET row_count = ? WHERE id = ?", (len(lead_ids), list_id))

    for lead_id in lead_ids:
        db.execute("INSERT OR IGNORE INTO list_leads (list_id, lead_id) VALUES (?, ?)", (list_id, lead_id))

    db.commit()
    flash(f"Imported {len(lead_ids)} leads into '{list_name}'", "success")
    return redirect(url_for("lists.detail", list_id=list_id))


@bp.route("/lists/<int:list_id>")
def detail(list_id):
    lst = query_db("SELECT * FROM lists WHERE id = ?", (list_id,), one=True)
    if not lst:
        flash("List not found", "error")
        return redirect(url_for("lists.index"))

    leads = query_db(
        """SELECT l.* FROM leads l
           JOIN list_leads ll ON l.id = ll.lead_id
           WHERE ll.list_id = ?
           ORDER BY CAST(l.rank AS INTEGER)""",
        (list_id,),
    )

    return render_template("lists/detail.html", lst=lst, leads=leads)


@bp.route("/lists/<int:list_id>/delete", methods=["POST"])
def delete(list_id):
    lst = query_db("SELECT * FROM lists WHERE id = ?", (list_id,), one=True)
    if not lst:
        flash("List not found", "error")
        return redirect(url_for("lists.index"))

    db = get_db()
    db.execute("DELETE FROM list_leads WHERE list_id = ?", (list_id,))
    db.execute("DELETE FROM lists WHERE id = ?", (list_id,))
    db.commit()

    flash(f"Deleted list '{lst['name']}'", "success")
    return redirect(url_for("lists.index"))


@bp.route("/lists/<int:list_id>/enrich", methods=["POST"])
def enrich(list_id):
    lst = query_db("SELECT * FROM lists WHERE id = ?", (list_id,), one=True)
    if not lst:
        flash("List not found", "error")
        return redirect(url_for("lists.index"))

    # Allow enrichment for 'none', 'error', and 'complete' (re-enrich)
    if lst["enrichment_status"] not in ("none", "error", "complete"):
        flash("Enrichment is already in progress", "info")
        return redirect(url_for("lists.detail", list_id=list_id))

    db = get_db()
    db.execute("UPDATE lists SET enrichment_status = 'enriching_urls' WHERE id = ?", (list_id,))
    db.commit()

    # Run enrichment in background thread
    app = current_app._get_current_object()

    def run_enrichment():
        from enrichment.pipeline import enrich_list
        with app.app_context():
            enrich_list(list_id)

    thread = threading.Thread(target=run_enrichment, daemon=True)
    thread.start()

    flash("Enrichment started. Progress will update automatically.", "info")
    return redirect(url_for("lists.detail", list_id=list_id))
